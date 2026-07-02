"""
Step 16: README

Regenerates the README sheet (sheet 1) with:
- Updated title row reflecting current sheet count and total data rows.
- Per-sheet line in the "Workbook structure" section updated with live counts.
- A "Last alignment" changelog entry appended at the bottom recording the
  current branch, ISO date, and a delta summary for sheets that changed.

Reads the canonical counts directly from the workbook (after steps 01-15
have run), so this step is idempotent against whatever state the workbook
is in.
"""
from __future__ import annotations

import datetime
import subprocess
from openpyxl.styles import Font, Alignment
from _common import ROOT, open_wb, save_wb


# Reference counts from the last known aligned commit (from original README)
PREVIOUS = {
    "Screens": 132,
    "Screen Actions": 1117,
    "Click Handlers": 1317,
    "Navigation Map": 192,
    "Components": 207,
    "Hooks": 25,
    "Client Library": 36,
    "Client Data & Tours": 8,
    "React Query Cache Keys": 388,
    "Client Configs": 10,
    "Toasts": 568,
    "App Entry": 8,
    "Data Sources": 1225,
    "Server Handlers": 1444,
    "Server Departments": 23,
    "Server Services": 52,
    "Server Repositories": 14,
    "Server Lib & Helpers": 72,
    "Server Policies & Infra": 5,
    "Middleware": 15,
    "Bootstrap & Schedulers": 31,
    "Background Jobs": 8,
    "Microsoft & Secrets": 3,
    "Database Tables": 280,
    "Table Relations": 518,
    "Database Indexes": 70,
    "SQL Migrations": 193,
    "Role & Permission Matrix": 648,
    "Feature Flags": 43,
    "Status Enums": 153,
    "Shared Business Logic": 22,
    "API Contracts & Validators": 4,
    "Environment Variables": 44,
    "Build Config": 22,
    "Dev Scripts": 26,
    "Tests": 192,
    "QA Infrastructure": 211,
    "Documentation": 292,
    "Skills": 4,
    "Public Assets": 9,
    "Attached Assets": 951,
    "Misc Artifacts": 11,
}

# Short descriptions for each sheet (kept to preserve the original README flavour)
SHEET_DESCRIPTIONS = {
    "Screens": "every screen.",
    "Screen Actions": "every button/field/filter/tab/link/API call on every screen.",
    "Click Handlers": "inline onClick={...} body of every clickable element.",
    "Navigation Map": "from→to navigation edges.",
    "Components": "React components in client/src/components.",
    "Hooks": "custom hooks in client/src/hooks.",
    "Client Library": "utility modules in client/src/lib (queryClient, api, errors, sanitize, idempotency, etc.).",
    "Client Data & Tours": "screen tours, walkthroughs, data/ folder CSVs, seed archives.",
    "React Query Cache Keys": "every queryKey first segment.",
    "Client Configs": "page-registry, app-navigation, role-dashboard, etc.",
    "Toasts": "every user-facing toast message.",
    "App Entry": "App.tsx, main.tsx, index.html, index.css, server/index.ts, db.ts, storage.ts, routes.ts.",
    "Data Sources": "every Express endpoint.",
    "Server Handlers": "endpoint→file→tables.",
    "Server Departments": "department route files.",
    "Server Services": "service modules.",
    "Server Repositories": "repository modules.",
    "Server Lib & Helpers": "helpers / utilities.",
    "Server Policies & Infra": "finance policy, write authority, bridge writer, import conflict policy, route ownership.",
    "Middleware": "Express middleware.",
    "Bootstrap & Schedulers": "startup / scheduler modules.",
    "Background Jobs": "long-running / scheduled tasks.",
    "Microsoft & Secrets": "MS presence, MS token manager, secrets vault.",
    "Database Tables": "every Drizzle table (shared/schema/**).",
    "Table Relations": "foreign-key edges between tables.",
    "Database Indexes": "every index defined in Drizzle schemas.",
    "SQL Migrations": "every SQL migration in migrations/.",
    "Role & Permission Matrix": "per-permission-entity → allowed-roles matrix.",
    "Feature Flags": "every flag from shared/feature-flags.ts (ROLLOUT_FEATURE_FLAGS).",
    "Status Enums": "pgEnum declarations and TS const-tuple enums.",
    "Shared Business Logic": "modules in shared/ outside of schema definitions.",
    "API Contracts & Validators": "zod / shared schemas used for request/response validation.",
    "Environment Variables": "every process.env.X reference.",
    "Build Config": "package.json, vite, drizzle, tsconfig, replit, .env examples, prettierrc, etc.",
    "Dev Scripts": "script/ + scripts/ folders (build, qa-report, route smoke, backfills, checks).",
    "Tests": "unit / integration test files.",
    "QA Infrastructure": "qa/ folder: playwright.config, vitest.config, release-gate, reports, tests.",
    "Documentation": "every .md file in the repo (excluding .agents/skills/).",
    "Skills": "agent skills under .agents/skills/.",
    "Public Assets": "static assets.",
    "Attached Assets": "test-fixture tracker workbooks under qa/fixtures/trackers/ (Excel).",
    "Misc Artifacts": "tsc logs, DB backup dumps, Excel inspectors, test-results logs, legacy folders.",
}


def get_branch():
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            cwd=ROOT, stderr=subprocess.DEVNULL,
        ).decode().strip()
    except Exception:
        return "unknown"


def section_header(title):
    return title


def build_readme_text(wb) -> list[str]:
    # Compute current counts per sheet
    counts = {}
    for name in wb.sheetnames:
        if name == "README":
            continue
        ws = wb[name]
        counts[name] = max(0, ws.max_row - 1)

    total = sum(counts.values())
    n_sheets = len(wb.sheetnames)

    lines: list[str] = []
    lines.append(f"Emergent Energy — System Map ({n_sheets} sheets, {total:,} rows)")
    lines.append(
        "The complete map of the Emergent Energy repository. Every screen, every button, "
        "every onClick, every API, every database table, every permission, every component, "
        "every hook, every library utility, every tour, every server service / repository / "
        "helper / policy / bridge, every scheduler, every middleware, every migration, every "
        "index, every feature flag, every cache key, every toast, every test, every doc, every "
        "build config, every dev script, every QA artifact, and every source asset file."
    )
    lines.append("")
    lines.append(f"Workbook structure ({n_sheets} sheets)")
    lines.append("")

    # Front-end block
    frontend = ["Screens", "Screen Actions", "Click Handlers", "Navigation Map",
                "Components", "Hooks", "Client Library", "Client Data & Tours",
                "React Query Cache Keys", "Client Configs", "Toasts", "App Entry"]
    backend = ["Data Sources", "Server Handlers", "Server Departments", "Server Services",
               "Server Repositories", "Server Lib & Helpers", "Server Policies & Infra",
               "Middleware", "Bootstrap & Schedulers", "Background Jobs", "Microsoft & Secrets"]
    database = ["Database Tables", "Table Relations", "Database Indexes", "SQL Migrations"]
    cross = ["Role & Permission Matrix", "Feature Flags", "Status Enums",
             "Shared Business Logic", "API Contracts & Validators", "Environment Variables"]
    other = ["Build Config", "Dev Scripts", "Tests", "QA Infrastructure",
             "Documentation", "Skills", "Public Assets", "Attached Assets", "Misc Artifacts"]

    def block(title, names, start_num):
        lines.append(f"{title} ({len(names)} sheets)")
        for i, name in enumerate(names, start=start_num):
            n = counts.get(name, 0)
            desc = SHEET_DESCRIPTIONS.get(name, "")
            prev = PREVIOUS.get(name)
            delta = ""
            if prev is not None:
                d = n - prev
                if d != 0:
                    delta = f" [{'+' if d > 0 else ''}{d} since last alignment]"
            lines.append(f" {i:>2}. {name} ({n:,}){' ' * max(1, 20 - len(name) - len(str(n)))}— {desc}{delta}")
        lines.append("")

    lines.append(" 1. README")
    block("FRONTEND", frontend, 2)
    block("BACKEND APIS & LOGIC", backend, 2 + len(frontend))
    block("DATABASE", database, 2 + len(frontend) + len(backend))
    block("CROSS-CUTTING", cross, 2 + len(frontend) + len(backend) + len(database))
    block("BUILD / TESTS / DOCS / SCRIPTS / ASSETS", other,
          2 + len(frontend) + len(backend) + len(database) + len(cross))

    lines.append(f"Total across all {n_sheets} sheets: {total:,} data rows.")
    lines.append("")
    lines.append("How to use it")
    lines.append(
        "All data sheets have a frozen header + AutoFilter. Sample queries:"
    )
    lines.append("  • 'Which dev scripts do we run before a release?' → Dev Scripts tab, filter purpose for 'release' / 'smoke' / 'qa-report'.")
    lines.append("  • 'What Excel source files does the app read from?' → Attached Assets tab, filter Extension = xlsx/xlsm.")
    lines.append("  • 'What finance policy rules does the server enforce?' → Server Policies & Infra tab, open 'finance-policy'.")
    lines.append("  • 'Which onClick body calls the API on the cashflow page?' → Click Handlers tab, filter Screen = cashflow.")
    lines.append("  • 'What client-side utility formats money?' → Client Library tab, search 'safeMoney'.")
    lines.append("  • 'Where are the screen tours for first-time users defined?' → Client Data & Tours tab, open 'screen-tours'.")
    lines.append("  • 'Which QA tests cover the auth routes?' → QA Infrastructure tab, filter 'auth-routes'.")
    lines.append("")
    lines.append("Source of truth")
    lines.append(
        "Everything in this workbook was generated from real source files in the repository. "
        "The refresh scripts live under scripts/workbook_refresh/ (step01..step16). "
        "Run them in order (or individually) whenever the repo structure changes."
    )
    lines.append(f"Branch: {get_branch()}")
    lines.append("")
    lines.append("Alignment notes")
    today = datetime.date.today().isoformat()
    lines.append(
        f"• {today}: full refresh against current repo. Key deltas vs the previous alignment:"
    )
    deltas = []
    for name, prev in PREVIOUS.items():
        cur = counts.get(name, 0)
        if cur != prev:
            deltas.append((name, prev, cur))
    # Sort by absolute delta desc
    deltas.sort(key=lambda x: abs(x[2] - x[1]), reverse=True)
    for name, prev, cur in deltas:
        d = cur - prev
        lines.append(f"    - {name}: {prev} → {cur} ({'+' if d > 0 else ''}{d})")
    lines.append("")
    lines.append(
        "• Narrative auto-generated sheets (Screen Actions, Click Handlers, Navigation Map, "
        "Toasts, React Query Cache Keys, Role & Permission Matrix) were NOT regenerated in "
        "this alignment — they still reflect the state at the last full autogen run. Re-run "
        "the legacy scripts (scripts/extract_page_actions.py, scripts/autogen_rows.py) to "
        "refresh them."
    )

    return lines


def main():
    wb = open_wb()
    ws = wb["README"]

    # Clear existing content
    if ws.max_row >= 1:
        ws.delete_rows(1, ws.max_row)

    lines = build_readme_text(wb)

    for i, line in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=line)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if i == 1:
            c.font = Font(bold=True, size=14)
        elif line and not line.startswith(" ") and not line.startswith("•") and not line.startswith("    "):
            # Section headers are non-indented single-line non-bullet entries
            if line.endswith(")") or line in ("How to use it", "Source of truth", "Alignment notes"):
                c.font = Font(bold=True, size=11)

    ws.column_dimensions["A"].width = 120
    save_wb(wb)
    print(f"README refreshed: {len(lines)} lines")


if __name__ == "__main__":
    main()
