"""
Shared helpers for the workbook refresh steps under scripts/workbook_refresh/.

Each step opens the workbook, rewrites one sheet's data region (row 2+),
preserves the column schema (row 1 headers) and re-applies the auto-filter.

Human-authored columns (e.g. plain-English descriptions) are preserved from
the existing workbook by keying on a stable identifier (SQL table name, URL,
path, etc.).
"""
from __future__ import annotations

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
XLSX = os.path.join(ROOT, "docs", "workbook", "emergent-energy-system-map.xlsx")
WRAP = Alignment(wrap_text=True, vertical="top")


def open_wb():
    return load_workbook(XLSX)


def save_wb(wb):
    wb.save(XLSX)


def snapshot(ws, key_col: int, value_cols: dict[int, str]) -> dict:
    """
    Read existing rows and snapshot columns of interest.
    key_col: 1-based column index used as the dict key.
    value_cols: { col_index: field_name } - columns to snapshot.
    Returns: { key_value: { field_name: cell_value_or_empty_string } }
    """
    out = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=key_col).value
        if not key:
            continue
        out[key] = {name: (ws.cell(row=r, column=col).value or "") for col, name in value_cols.items()}
    return out


def clear_data(ws):
    """
    Delete row 2+ (keeps header row).
    Uses delete_rows so ws.max_row actually shrinks — setting cells to None
    leaves phantom trailing rows.
    """
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)


def write_rows(ws, rows: list[list]):
    """Write rows starting at row 2, wrapping text. Rows are lists aligned with the header."""
    for i, row in enumerate(rows, start=2):
        for j, v in enumerate(row, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.alignment = WRAP


def apply_autofilter(ws, last_row: int):
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{last_row}"


def walk_files(relative_root: str, exts: tuple[str, ...]):
    """Yield repo-relative paths under ROOT/relative_root matching any extension in exts."""
    base = os.path.join(ROOT, relative_root)
    for dirpath, dirnames, filenames in os.walk(base):
        dirnames[:] = [d for d in dirnames if d not in ("node_modules", ".git", "dist", "build", ".next")]
        for f in filenames:
            if exts and not f.endswith(exts):
                continue
            full = os.path.join(dirpath, f)
            yield os.path.relpath(full, ROOT)
