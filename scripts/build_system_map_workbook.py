"""
Build the business-friendly Excel workbook for the Emergent Energy web app.

Source data:
- /tmp/tables.txt   - 280 Drizzle tables (JS name + SQL name + source file)
- /tmp/all_routes.txt - 1225 unique (METHOD, PATH) pairs
- /home/user/Emergent-Energy-Web-App/client/src/config/page-registry.ts - 128 page entries

Output:
- /home/user/Emergent-Energy-Web-App/docs/workbook/emergent-energy-system-map.xlsx
"""
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = "/home/user/Emergent-Energy-Web-App"
OUT = f"{ROOT}/docs/workbook/emergent-energy-system-map.xlsx"

# ============================================================================
# STYLES
# ============================================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # deep blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")  # light blue
SECTION_FONT = Font(bold=True, color="1F4E78", size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def style_header(ws, row_idx, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ============================================================================
# DATA LOAD: TABLES
# ============================================================================
def load_tables():
    rows = []
    with open("/tmp/tables.txt") as f:
        for line in f:
            line = line.strip()
            m = re.match(r'export const (\w+) = pgTable\(\s*"?([\w_]+)?"?', line)
            if not m:
                continue
            js = m.group(1)
            sql = m.group(2) or ""
            fm = re.search(r"\(([^)]+)\)$", line)
            source = fm.group(1) if fm else ""
            rows.append({"js": js, "sql": sql, "source": source})
    return rows

# ============================================================================
# DATA LOAD: ROUTES
# ============================================================================
def load_routes():
    rows = []
    with open("/tmp/all_routes.txt") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split(None, 1)
            if len(parts) != 2:
                continue
            method, path = parts
            rows.append({"method": method, "path": path})
    return rows

# ============================================================================
# DATA LOAD: PAGE REGISTRY
# ============================================================================
def load_pages():
    with open(f"{ROOT}/client/src/config/page-registry.ts") as f:
        content = f.read()
    # Crude but effective: capture each { ... } block inside PAGE_REGISTRY
    entries = []
    # Match a single-line entry. The registry is mostly single-line entries.
    line_re = re.compile(r'^\s*\{\s*id:\s*"([^"]+)",\s*path:\s*"([^"]+)",\s*label:\s*"([^"]+)"(.*?)\},?\s*$')
    for line in content.splitlines():
        m = line_re.match(line)
        if not m:
            continue
        rest = m.group(4)
        nav = re.search(r'navGroup:\s*"([^"]+)"', rest)
        perm = re.search(r'permissionEntity:\s*"([^"]+)"', rest)
        rck = re.search(r'routeComponentKey:\s*"([^"]+)"', rest)
        redirect = re.search(r'redirectTo:\s*"([^"]+)"', rest)
        entries.append({
            "id": m.group(1),
            "path": m.group(2),
            "label": m.group(3),
            "navGroup": nav.group(1) if nav else "",
            "permissionEntity": perm.group(1) if perm else "",
            "routeComponentKey": rck.group(1) if rck else "",
            "redirectTo": redirect.group(1) if redirect else "",
            "isAlias": 'type: "alias"' in rest or redirect is not None,
        })
    return entries

TABLES = load_tables()
ROUTES = load_routes()
PAGES = load_pages()

print(f"Loaded {len(TABLES)} tables, {len(ROUTES)} routes, {len(PAGES)} registry entries")
