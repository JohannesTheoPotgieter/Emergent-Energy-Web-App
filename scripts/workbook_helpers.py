"""
Helpers used by the incremental workbook builder steps.

Each step opens the xlsx, calls one of these helpers, saves.
"""
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

XLSX = "/home/user/Emergent-Energy-Web-App/docs/workbook/emergent-energy-system-map.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, color="1F4E78", size=12)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def open_wb():
    return load_workbook(XLSX)


def save_wb(wb):
    wb.save(XLSX)


def write_header_row(ws, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER


def apply_widths(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def apply_autofilter(ws):
    from openpyxl.utils import get_column_letter
    max_col = ws.max_column
    max_row = ws.max_row
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


def fill_table_descriptions(ws, descriptions: dict):
    """
    descriptions: { sql_name: (friendly_name, description) }
    Walks the existing Database Tables sheet rows and fills in columns C (friendly)
    and D (description) for any matching row by SQL name in column B.
    Returns the number of rows updated.
    """
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        sql = ws.cell(row=row_idx, column=2).value
        if not sql:
            continue
        if sql in descriptions:
            friendly, desc = descriptions[sql]
            c_friendly = ws.cell(row=row_idx, column=3, value=friendly)
            c_friendly.alignment = WRAP
            c_desc = ws.cell(row=row_idx, column=4, value=desc)
            c_desc.alignment = WRAP
            updated += 1
    return updated


def append_rows_to_sheet(ws, rows: list, start_col: int = 1):
    """Append each tuple/list as a row starting after the last non-empty row."""
    start_row = ws.max_row + 1 if ws.max_row else 1
    for i, row in enumerate(rows):
        for j, v in enumerate(row, start=start_col):
            c = ws.cell(row=start_row + i, column=j, value=v)
            c.alignment = WRAP
    return len(rows)


def fill_screen_descriptions(ws, descriptions: dict):
    """
    descriptions: { url_path: (friendly_purpose, who_uses_it, key_data_shown) }
    Walks the Screens sheet and fills cols D (Purpose), E (Who Uses It), F (Key Data)
    where the URL path (column B) matches.
    """
    updated = 0
    for row_idx in range(2, ws.max_row + 1):
        url = ws.cell(row=row_idx, column=2).value
        if not url:
            continue
        if url in descriptions:
            purpose, who, data = descriptions[url]
            for col, val in [(4, purpose), (5, who), (6, data)]:
                c = ws.cell(row=row_idx, column=col, value=val)
                c.alignment = WRAP
            updated += 1
    return updated
